"""Rendering a :class:`~app.exporters.report.Report` to CSV and XLSX bytes.

Both writers put every cell through :mod:`app.exporters.cells`, so a material
named ``=cmd|'/c calc'!A1`` leaves as inert text in either format.

CSV holds one table per file by convention, but a selection report is only
trustworthy whole — the funnel means nothing without the constraints that
produced it. So the CSV carries every section in one file, separated by a blank
line and a title row. It reads fine in a spreadsheet and stays a single
attachment.
"""

from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.exporters.cells import safe_number, safe_text
from app.exporters.report import Report, Sheet

# Excel rejects : \ / ? * [ ] in sheet names and caps them at 31 characters.
_INVALID_SHEET_CHARS = str.maketrans({c: "-" for c in ":\\/?*[]"})
_MAX_SHEET_NAME = 31

# UTF-8 BOM: without it Excel on Windows opens a UTF-8 CSV as Latin-1 and turns
# "Módulo" into "MÃ³dulo". The report is in Portuguese, so this is not optional.
_BOM = "﻿"


def _cell(value: object) -> object:
    """Escape a value for a spreadsheet, keeping real numbers numeric."""
    if isinstance(value, bool):
        return safe_text("sim" if value else "não")
    if isinstance(value, (int, float)):
        return safe_number(value)
    return safe_text(value)


def to_csv(report: Report) -> str:
    """Render the whole report as one CSV document.

    ``\\r\\n`` line endings and QUOTE_MINIMAL match what spreadsheets expect;
    the writer quotes any field containing a separator or newline itself.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([safe_text(report.title)])
    if report.subtitle:
        writer.writerow([safe_text(report.subtitle)])
    for notice in report.notices:
        writer.writerow([safe_text(notice)])

    for sheet in report.sheets:
        writer.writerow([])
        writer.writerow([safe_text(sheet.name)])
        for note in sheet.notes:
            writer.writerow([safe_text(note)])
        if sheet.header:
            writer.writerow([safe_text(column) for column in sheet.header])
        for row in sheet.rows:
            writer.writerow([_cell(value) for value in row])

    return _BOM + buffer.getvalue()


def _unique_sheet_name(raw: str, taken: set[str]) -> str:
    """Excel-legal, unique worksheet name."""
    name = raw.translate(_INVALID_SHEET_CHARS)[:_MAX_SHEET_NAME] or "Planilha"
    if name not in taken:
        return name
    for suffix in range(2, 100):
        candidate = f"{name[: _MAX_SHEET_NAME - len(str(suffix)) - 1]} {suffix}"
        if candidate not in taken:
            return candidate
    return name[: _MAX_SHEET_NAME - 4] + " ..."  # pragma: no cover - absurd input


def _write_sheet(worksheet, sheet: Sheet) -> None:
    bold = Font(bold=True)
    for note in sheet.notes:
        worksheet.append([safe_text(note)])
    if sheet.header:
        worksheet.append([safe_text(column) for column in sheet.header])
        for cell in worksheet[worksheet.max_row]:
            cell.font = bold
    for row in sheet.rows:
        worksheet.append([_cell(value) for value in row])

    # Width from the longest value in each column, so the trail is readable
    # without the reader having to resize anything.
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 60)


def to_xlsx(report: Report) -> bytes:
    """Render the report as a workbook, one worksheet per section plus a cover."""
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Aviso"
    cover.append([safe_text(report.title)])
    cover["A1"].font = Font(bold=True, size=14)
    if report.subtitle:
        cover.append([safe_text(report.subtitle)])
    cover.append([])
    for notice in report.notices:
        cover.append([safe_text(notice)])
        cover[f"A{cover.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.column_dimensions["A"].width = 110

    taken = {cover.title}
    for sheet in report.sheets:
        name = _unique_sheet_name(sheet.name, taken)
        taken.add(name)
        _write_sheet(workbook.create_sheet(title=name), sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
