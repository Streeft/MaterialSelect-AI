"""Tests for the export layer.

The central case is injection, and it takes a different shape per format. A
material name is free text that can arrive through the import wizard or the
manual form: unescaped in a CSV it runs when the file is opened, and unescaped
in the printable HTML it runs when the page is viewed. Both tests therefore
store a hostile name through the real API and read it back out of a real
export, rather than asserting against a hand-built string.
"""

from __future__ import annotations

import csv
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.exporters.cells import format_number, is_dangerous, safe_number, safe_text
from app.exporters.html import to_html
from app.exporters.report import (
    DEMO_DATA_NOTICE,
    LIMITATION_NOTICE,
    Report,
    Sheet,
    standard_notices,
)
from app.exporters.spreadsheet import to_csv, to_xlsx
from app.routers.exports import HTML_CSP

HOSTILE_NAMES = [
    "=cmd|'/c calc'!A1",
    "+1+1",
    "-2+3",
    "@SUM(A1:A9)",
    '=HYPERLINK("http://exemplo","clique")',
]


class TestCellSafety:
    @pytest.mark.parametrize("name", HOSTILE_NAMES)
    def test_formula_prefixes_are_detected(self, name: str) -> None:
        assert is_dangerous(name)

    @pytest.mark.parametrize("name", HOSTILE_NAMES)
    def test_dangerous_text_is_marked_as_literal(self, name: str) -> None:
        escaped = safe_text(name)
        assert escaped.startswith("'")
        # Escaping is visible, not destructive: the original survives intact.
        assert escaped[1:] == name

    def test_ordinary_text_is_untouched(self) -> None:
        assert safe_text("Aço Demo B") == "Aço Demo B"

    def test_none_becomes_an_empty_cell(self) -> None:
        assert safe_text(None) == ""

    def test_negative_numbers_stay_numeric(self) -> None:
        # A numeric cell is never parsed as a formula, and escaping it would
        # make arithmetic in the exported sheet impossible.
        assert safe_number(-40.5) == -40.5

    def test_missing_number_is_a_word_not_a_zero(self) -> None:
        assert format_number(None) == "ausente"
        assert format_number(None, missing="indefinido") == "indefinido"

    def test_integers_render_without_a_decimal_tail(self) -> None:
        assert format_number(2700.0) == "2700"

    def test_conversion_residue_does_not_reach_the_reader(self) -> None:
        # 3.9 g/cm**3 normalizes to this exact double. Printing all seventeen
        # digits would claim a precision the measurement never had.
        assert format_number(3.9 * 1000) == "3900"
        assert format_number(0.1 + 0.2) == "0.3"

    def test_real_precision_survives(self) -> None:
        # Rounding is a reporting choice, not a licence to round the datum away.
        assert format_number(1234.56789012) == "1234.56789012"
        assert format_number(2.5e-05) == "2.5e-05"
        assert format_number(-40.5) == "-40.5"


class TestNotices:
    def test_limitation_notice_is_always_present(self) -> None:
        assert LIMITATION_NOTICE in standard_notices(includes_demo_data=False)

    def test_demo_warning_comes_first_when_it_applies(self) -> None:
        notices = standard_notices(includes_demo_data=True)
        assert notices[0] == DEMO_DATA_NOTICE

    def test_no_demo_warning_when_no_demo_data(self) -> None:
        assert DEMO_DATA_NOTICE not in standard_notices(includes_demo_data=False)


def _report() -> Report:
    return Report(
        title="Relatório de teste",
        subtitle="subtítulo",
        notices=standard_notices(includes_demo_data=True),
        sheets=[
            Sheet(
                name="Dados",
                header=["Material", "Valor"],
                rows=[["=SOMA(A1)", 2700.0], ["Aço Demo B", None]],
                notes=["uma observação"],
            )
        ],
    )


class TestCsvRendering:
    def test_hostile_cell_is_escaped(self) -> None:
        assert "'=SOMA(A1)" in to_csv(_report())

    def test_notices_are_written_before_the_data(self) -> None:
        text = to_csv(_report())
        assert text.index(LIMITATION_NOTICE) < text.index("Material")

    def test_starts_with_a_bom_so_excel_reads_utf8(self) -> None:
        # Without it Excel on Windows renders "Relatório" as "RelatÃ³rio".
        assert to_csv(_report()).startswith("﻿")

    def test_is_parseable_and_keeps_the_header(self) -> None:
        rows = list(csv.reader(io.StringIO(to_csv(_report()).lstrip("﻿"))))
        assert ["Material", "Valor"] in rows


class TestXlsxRendering:
    def test_workbook_has_a_notice_cover_plus_each_sheet(self) -> None:
        workbook = load_workbook(io.BytesIO(to_xlsx(_report())))
        assert workbook.sheetnames == ["Aviso", "Dados"]

    def test_hostile_cell_is_escaped_in_xlsx_too(self) -> None:
        workbook = load_workbook(io.BytesIO(to_xlsx(_report())))
        values = [cell.value for row in workbook["Dados"].iter_rows() for cell in row]
        assert "'=SOMA(A1)" in values
        assert "=SOMA(A1)" not in values

    def test_numbers_stay_numbers(self) -> None:
        workbook = load_workbook(io.BytesIO(to_xlsx(_report())))
        values = [cell.value for row in workbook["Dados"].iter_rows() for cell in row]
        assert 2700.0 in values

    def test_sheet_names_are_made_excel_legal(self) -> None:
        report = _report()
        report.sheets.append(Sheet(name="Inválido: nome/com*chars", header=["a"], rows=[]))
        workbook = load_workbook(io.BytesIO(to_xlsx(report)))
        assert all(not set(name) & set(":\\/?*[]") for name in workbook.sheetnames)


class TestHtmlRendering:
    """The printable report.

    Its risk profile is not the spreadsheet's. Nothing here can run a formula,
    but the document is served as markup on the API's own origin, so a material
    name is a script-injection vector unless every value is escaped.
    """

    def test_markup_in_a_value_cannot_become_markup(self) -> None:
        report = _report()
        report.sheets[0].rows.append(["<script>alert(1)</script>", 1.0])
        rendered = to_html(report)
        assert "<script>" not in rendered
        assert "&lt;script&gt;alert(1)&lt;/script&gt;" in rendered

    def test_markup_in_a_heading_or_note_is_escaped_too(self) -> None:
        report = _report()
        report.sheets[0].name = "<img onerror=x>"
        report.sheets[0].notes = ["<b>nota</b>"]
        report.sheets[0].header = ["<i>coluna</i>", "Valor"]
        rendered = to_html(report)
        assert "<img" not in rendered
        assert "<b>nota</b>" not in rendered
        assert "<i>coluna</i>" not in rendered

    def test_the_spreadsheet_apostrophe_does_not_leak_into_html(self) -> None:
        # A leading "=" is inert in HTML; prefixing an apostrophe here would
        # corrupt the exported value for no gain.
        rendered = to_html(_report())
        assert "&#x27;=SOMA" not in rendered
        assert "'=SOMA" not in rendered
        assert "=SOMA(A1)" in rendered

    def test_missing_value_is_a_word_not_an_empty_cell(self) -> None:
        # The row fixture carries a None; a blank cell would read as zero.
        assert "<td>ausente</td>" in to_html(_report())

    def test_notices_appear_before_the_data(self) -> None:
        rendered = to_html(_report())
        assert rendered.index(LIMITATION_NOTICE) < rendered.index("Material")

    def test_demo_warning_is_singled_out(self) -> None:
        # Assert on the class being *applied*, not on the stylesheet that
        # defines it — the rule is present in every document either way.
        assert 'class="notice notice-demo"' in to_html(_report())
        without_demo = Report(
            title="t",
            subtitle="",
            notices=standard_notices(includes_demo_data=False),
            sheets=[],
        )
        assert 'class="notice notice-demo"' not in to_html(without_demo)

    def test_document_is_self_contained(self) -> None:
        # It has to render identically offline, from an email attachment,
        # years from now. Any external reference breaks that.
        rendered = to_html(_report())
        assert "<script" not in rendered
        assert "<link" not in rendered
        assert "src=" not in rendered

    def test_carries_print_rules(self) -> None:
        rendered = to_html(_report())
        assert "@media print" in rendered
        # Repeating the header on each printed page is what keeps a table
        # readable once it spills past the first one.
        assert "table-header-group" in rendered

    def test_render_is_deterministic(self) -> None:
        # No timestamp, no ordering wobble: the same report is the same bytes.
        assert to_html(_report()) == to_html(_report())

    def test_sections_are_numbered_for_citation(self) -> None:
        assert "<h2>1. Dados</h2>" in to_html(_report())

    def test_responsible_is_absent_by_default(self) -> None:
        assert '<p class="responsible">' not in to_html(_report())

    def test_responsible_name_is_escaped(self) -> None:
        report = _report()
        report.responsible = "<b>Fulano</b>"
        rendered = to_html(report)
        assert "<b>Fulano</b>" not in rendered
        assert "Responsável técnico: &lt;b&gt;Fulano&lt;/b&gt;" in rendered

    def test_figure_is_embedded_as_trusted_markup(self) -> None:
        report = _report()
        report.figure = '<svg role="img"><title>x</title></svg>'
        rendered = to_html(report)
        # Not escaped: this is our own SVG output, not reader-supplied text.
        assert '<div class="figure"><svg role="img">' in rendered

    def test_narrative_paragraphs_are_escaped_and_numbered_after_the_sheets(self) -> None:
        report = _report()
        report.narrative = ["<script>alert(1)</script>", "Parágrafo normal."]
        rendered = to_html(report)
        assert "<h2>2. Interpretação técnica (IA)</h2>" in rendered
        assert "<script>" not in rendered
        assert "Parágrafo normal." in rendered

    def test_narrative_absence_is_declared_not_omitted(self) -> None:
        report = _report()
        report.narrative_note = "Interpretação por IA não disponível: camada desligada."
        rendered = to_html(report)
        assert "Interpretação técnica (IA)" in rendered
        assert "Interpretação por IA não disponível" in rendered

    def test_no_narrative_section_when_neither_field_is_set(self) -> None:
        assert "Interpretação técnica" not in to_html(_report())


class TestHtmlEndpoint:
    def test_catalogue_html_opens_inline_rather_than_downloading(self, client: TestClient) -> None:
        response = client.get("/api/exports/catalogo.html")
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith("text/html")
        # Inline is the point: the browser renders it so the user can print it.
        assert response.headers["content-disposition"].startswith("inline")

    def test_html_is_served_under_a_no_execution_policy(self, client: TestClient) -> None:
        response = client.get("/api/exports/catalogo.html")
        assert response.headers["content-security-policy"] == HTML_CSP
        assert response.headers["x-content-type-options"] == "nosniff"

    def test_spreadsheet_formats_still_download(self, client: TestClient) -> None:
        for fmt in ("csv", "xlsx"):
            disposition = client.get(f"/api/exports/catalogo.{fmt}").headers["content-disposition"]
            assert disposition.startswith("attachment"), fmt

    def test_html_carries_the_mandatory_notices(self, client: TestClient) -> None:
        text = client.get("/api/exports/catalogo.html").text
        assert LIMITATION_NOTICE in text
        assert DEMO_DATA_NOTICE in text

    def test_a_hostile_material_name_cannot_escape_as_markup(self, client: TestClient) -> None:
        created = client.post(
            "/api/materials",
            json={
                "name": '<img src=x onerror="alert(1)">',
                "class_id": 1,
                "keywords": [],
                "values": [
                    {
                        "property_slug": "densidade",
                        "kind": "scalar",
                        "value": 1000.0,
                        "unit": "kg/m**3",
                        "data_quality": "ESTIMADO",
                    }
                ],
            },
        )
        assert created.status_code == 201, created.text

        text = client.get("/api/exports/catalogo.html").text
        assert "<img src=x" not in text
        assert "onerror" not in text or "&quot;" in text
        assert "&lt;img src=x" in text

    def test_unknown_study_is_404(self, client: TestClient) -> None:
        assert client.get("/api/exports/estudos/999999.html").status_code == 404

    def test_unsupported_format_is_still_rejected(self, client: TestClient) -> None:
        assert client.get("/api/exports/catalogo.pdf").status_code == 400


class TestCatalogueExport:
    def test_csv_export_succeeds(self, client: TestClient) -> None:
        response = client.get("/api/exports/catalogo.csv")
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith("text/csv")
        assert response.headers["x-content-type-options"] == "nosniff"
        assert "attachment" in response.headers["content-disposition"]

    def test_csv_carries_the_mandatory_notices(self, client: TestClient) -> None:
        text = client.get("/api/exports/catalogo.csv").text
        assert LIMITATION_NOTICE in text
        assert DEMO_DATA_NOTICE in text  # the seeded catalogue is demo data

    def test_missing_value_is_reported_as_absent(self, client: TestClient) -> None:
        # Cerâmica Demo D has an explicitly missing thermal conductivity.
        text = client.get("/api/exports/catalogo.csv").text
        assert "ausente" in text

    def test_xlsx_export_is_a_real_workbook(self, client: TestClient) -> None:
        response = client.get("/api/exports/catalogo.xlsx")
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert "Materiais" in workbook.sheetnames
        assert "Proveniência" in workbook.sheetnames

    def test_provenance_records_the_conversion_trail(self, client: TestClient) -> None:
        text = client.get("/api/exports/catalogo.csv").text
        assert "pint:GPa->Pa" in text

    def test_unsupported_format_is_rejected(self, client: TestClient) -> None:
        assert client.get("/api/exports/catalogo.pdf").status_code == 400

    def test_a_hostile_material_name_cannot_escape_as_a_formula(self, client: TestClient) -> None:
        created = client.post(
            "/api/materials",
            json={
                "name": "=cmd|'/c calc'!A1",
                "class_id": 1,
                "keywords": [],
                "values": [
                    {
                        "property_slug": "densidade",
                        "kind": "scalar",
                        "value": 1000.0,
                        "unit": "kg/m**3",
                        "data_quality": "ESTIMADO",
                    }
                ],
            },
        )
        assert created.status_code == 201, created.text

        text = client.get("/api/exports/catalogo.csv").text
        assert "'=cmd|'/c calc'!A1" in text or "\"'=cmd|'/c calc'!A1\"" in text
        # The unescaped form must not appear at the start of any field.
        rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))
        assert not any(
            isinstance(cell, str) and cell.startswith("=") for row in rows for cell in row
        )


def _exportable_study_id(client: TestClient) -> int:
    response = client.post(
        "/api/selection/studies",
        json={
            "name": "Estudo exportável",
            "function_text": "Viga em flexão",
            "objective_text": "Minimizar massa",
            "free_variables": ["espessura"],
            "combinator": "AND",
            "constraints": [
                {"operator": "gt", "property_slug": "modulo_young", "value": 1.0, "unit": "GPa"}
            ],
            "index": {
                "name": "Viga leve",
                "expression": "sqrt(modulo_young) / densidade",
                "goal": "maximize",
            },
            "normalization": "minmax",
            "criteria": [
                {"key": "__index__", "weight": 2.0},
                {"key": "densidade", "weight": 1.0},
            ],
        },
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


class TestStudyExport:
    def _study_id(self, client: TestClient) -> int:
        return _exportable_study_id(client)

    def test_report_contains_every_audit_section(self, client: TestClient) -> None:
        response = client.get(f"/api/exports/estudos/{self._study_id(client)}.xlsx")
        assert response.status_code == 200, response.text
        names = load_workbook(io.BytesIO(response.content)).sheetnames
        for expected in [
            "Aviso",
            "Problema",
            "Restrições e funil",
            "Candidatos",
            "Índice de desempenho",
            "Contribuições",
            "Excluídos por dado ausente",
            "Sensibilidade",
            "Proveniência",
        ]:
            assert expected in names, f"faltou a aba '{expected}' em {names}"

    def test_html_report_contains_every_audit_section(self, client: TestClient) -> None:
        # Same Report, different renderer: the printable version must not be a
        # reduced summary of what the spreadsheet carries.
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.html").text
        for expected in [
            "Problema",
            "Restrições e funil",
            "Candidatos",
            "Índice de desempenho",
            "Contribuições",
            "Excluídos por dado ausente",
            "Sensibilidade",
            "Proveniência",
        ]:
            assert expected in text, f"faltou a seção '{expected}'"

    def test_html_report_states_the_derived_dimension(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.html").text
        assert "Dimensão (derivada)" in text
        assert "[length]" in text

    def test_csv_report_states_the_derived_dimension(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.csv").text
        assert "Dimensão (derivada)" in text
        assert "[length]" in text

    def test_report_never_prints_a_raw_key_where_a_name_belongs(self, client: TestClient) -> None:
        """Slugs and "__index__" identify things; they are not words for a reader.

        Three surfaces used to leak them: the criterion column of the
        contributions table and the sensitivity scenarios (a saved study
        defaulted its label to the key), the excluded table, and the
        provenance row for a property a material has no entry for at all —
        which read the name off the value that was missing.
        """
        study_id = client.post(
            "/api/selection/studies",
            json={
                "name": "Estudo sem chaves cruas",
                "combinator": "AND",
                "constraints": [],
                "index": {
                    "name": "Viga leve",
                    "expression": "sqrt(modulo_young) / densidade",
                    "goal": "maximize",
                },
                "normalization": "minmax",
                "criteria": [
                    {"key": "__index__", "weight": 2.0},
                    # Three of the five demo materials have no yield strength,
                    # so this criterion produces both an exclusion row and a
                    # provenance row with no value behind it.
                    {"key": "limite_escoamento", "weight": 1.0},
                ],
            },
        ).json()["id"]

        text = client.get(f"/api/exports/estudos/{study_id}.html").text
        assert "__index__" not in text
        assert "Ênfase em Viga leve" in text
        # The slug appears nowhere; the property's name appears instead. The
        # index expression legitimately carries other slugs, so this asserts on
        # the one property that is not in it.
        assert "limite_escoamento" not in text
        assert "Limite de escoamento" in text

    def test_report_records_the_funnel(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.csv").text
        assert "Restrições e funil" in text
        assert "Módulo de Young" in text

    def test_report_carries_the_limitation_notice(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.csv").text
        assert LIMITATION_NOTICE in text

    def test_provenance_lists_the_properties_the_decision_used(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{self._study_id(client)}.csv").text
        provenance = text.split("Proveniência")[-1]
        assert "Densidade" in provenance
        assert "Módulo de Young" in provenance

    def test_filename_is_ascii_safe(self, client: TestClient) -> None:
        response = client.get(f"/api/exports/estudos/{self._study_id(client)}.csv")
        disposition = response.headers["content-disposition"]
        plain = disposition.split('filename="')[1].split('"')[0]
        assert plain.isascii()
        assert "filename*=UTF-8''" in disposition

    def test_unknown_study_is_404(self, client: TestClient) -> None:
        assert client.get("/api/exports/estudos/999999.csv").status_code == 404

    def test_report_describes_the_real_nested_constraint_tree(self, client: TestClient) -> None:
        """A nested study used to be described by ``study.combinator`` alone
        — the root group's own operator, presented as if it were the whole
        study's logic — with each subgroup an opaque "Subgrupo" funnel row
        showing nothing about what was inside it (D-41's laudo has to
        represent the study's real logic, not just its final numbers).

        This asserts the "Problema" sheet now spells out the real tree: the
        root OR, both nested AND groups, and a constraint from each — not
        just "OR" alone.
        """
        payload = {
            "name": "Estudo aninhado exportável",
            "root_group": {
                "operator": "OR",
                "constraints": [],
                "groups": [
                    {
                        "operator": "AND",
                        "constraints": [
                            {
                                "operator": "between",
                                "property_slug": "densidade",
                                "value_min": 2000,
                                "value_max": 3000,
                                "unit": "kg/m**3",
                            },
                        ],
                        "groups": [],
                    },
                    {
                        "operator": "AND",
                        "constraints": [
                            {
                                "operator": "between",
                                "property_slug": "modulo_young",
                                "value_min": 200,
                                "value_max": 250,
                                "unit": "GPa",
                            },
                        ],
                        "groups": [],
                    },
                ],
            },
            "normalization": "minmax",
            "criteria": [],
        }
        study_id = client.post("/api/selection/studies", json=payload).json()["id"]

        text = client.get(f"/api/exports/estudos/{study_id}.csv").text
        problem = text.split("Restrições e funil")[0]
        assert "OU(" in problem
        assert problem.count("E(") == 2
        assert "Densidade" in problem
        assert "Módulo de Young" in problem


class TestStudyLaudo:
    """The engineering report: a document distinct from the selection report.

    It carries the same audit sections, plus a ranking figure and (when the
    AI layer is on, which it is by default in tests — ``AI_PROVIDER=mock``)
    an interpretive narrative.
    """

    def test_laudo_contains_every_audit_section(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        for expected in [
            "Problema",
            "Restrições e funil",
            "Candidatos",
            "Índice de desempenho",
            "Contribuições",
            "Excluídos por dado ausente",
            "Sensibilidade",
            "Proveniência",
        ]:
            assert expected in text, f"faltou a seção '{expected}'"

    def test_laudo_title_differs_from_the_selection_report(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        assert "<h1>Laudo de engenharia" in text

    def test_laudo_embeds_a_ranking_figure(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        assert '<div class="figure">' in text
        assert 'role="img"' in text
        assert "Candidatos ranqueados" in text

    def test_laudo_includes_the_ai_narrative_by_default(self, client: TestClient) -> None:
        # AI_PROVIDER defaults to "mock" — deterministic, no network — so the
        # narrative section is populated out of the box.
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        assert "Interpretação técnica (IA)" in text
        assert "Interpretação por IA não disponível" not in text

    def test_laudo_declares_narrative_absence_when_ai_is_off(
        self, client: TestClient, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from app.services import ai_service

        monkeypatch.setattr(ai_service.default_settings, "ai_provider", "")
        response = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html")
        assert response.status_code == 200, response.text
        assert "Interpretação por IA não disponível" in response.text

    def test_laudo_declares_the_responsible_engineer_when_given(self, client: TestClient) -> None:
        study_id = _exportable_study_id(client)
        response = client.get(
            f"/api/exports/estudos/{study_id}/laudo.html", params={"responsavel": "Ana Engenheira"}
        )
        assert "Responsável técnico: Ana Engenheira" in response.text

    def test_laudo_omits_the_responsible_line_when_not_given(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        assert "Responsável técnico" not in text

    def test_a_hostile_responsible_name_cannot_escape_as_markup(self, client: TestClient) -> None:
        study_id = _exportable_study_id(client)
        response = client.get(
            f"/api/exports/estudos/{study_id}/laudo.html",
            params={"responsavel": "<img src=x onerror=alert(1)>"},
        )
        assert "<img src=x" not in response.text
        assert "&lt;img src=x" in response.text

    def test_laudo_is_served_inline_under_the_no_execution_policy(self, client: TestClient) -> None:
        response = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html")
        assert response.headers["content-disposition"].startswith("inline")
        assert response.headers["content-security-policy"] == HTML_CSP

    def test_laudo_carries_the_limitation_notice(self, client: TestClient) -> None:
        text = client.get(f"/api/exports/estudos/{_exportable_study_id(client)}/laudo.html").text
        assert LIMITATION_NOTICE in text

    def test_unknown_study_is_404(self, client: TestClient) -> None:
        assert client.get("/api/exports/estudos/999999/laudo.html").status_code == 404
