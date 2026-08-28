"""File readers for the import wizard (CSV, XLSX, JSON and SQLite).

All readers return the same shape — ``TabularData`` with headers and rows of
raw cell values — so the rest of the pipeline is format-agnostic. Design notes:

* CSV: encoding tried as UTF-8 (with BOM) first, then Latin-1; the delimiter is
  sniffed among ``;``, ``,`` and TAB (Brazilian spreadsheets commonly use ";").
* XLSX: read with ``openpyxl`` in ``data_only=True`` mode, so formula cells
  yield their cached values and never formula text. ``read_only=True`` bounds
  memory usage.
* JSON: parses a JSON array of flat objects, extracting headers from the union
  of object keys (in first-seen order) and filling missing keys with None.
* SQLite: extracts one table from a SQLite database; with no table name given,
  the first user table (excluding SQLite internal tables) is used.
* stdlib ``csv`` + ``openpyxl`` were chosen over pandas: every needed feature
  is covered without a heavyweight dependency (documented in docs/06-importacao.md).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from charset_normalizer import from_bytes
from openpyxl import load_workbook

from app.domain.errors import ValidationError

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".json", ".sqlite", ".db"}


@dataclass
class TabularData:
    """Headers plus raw data rows read from one sheet/file."""

    headers: list[str]
    rows: list[list[object]]
    sheet_names: list[str] = field(default_factory=list)
    sheet_name: str | None = None


def _decode_csv(data: bytes) -> str:
    # utf-8-sig first: strict, so it never falsely matches non-UTF-8 bytes.
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    # charset-normalizer scores candidate encodings by how "natural" the
    # decoded text looks (character frequency, mojibake detection) instead of
    # accepting the first encoding that merely doesn't raise — which is why
    # Latin-1 alone (every byte is valid Latin-1) used to mask cp1252/other
    # Windows encodings silently.
    best = from_bytes(data).best()
    if best is not None:
        return str(best)
    # Latin-1 never raises, so this is the final, always-successful fallback
    # for any byte sequence charset-normalizer scored too low to trust.
    try:
        return data.decode("latin-1")
    except UnicodeDecodeError as exc:
        raise ValidationError("Não foi possível decodificar o arquivo CSV (tente UTF-8).") from exc


def _sniff_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
    except csv.Error:
        # Fall back to the delimiter that appears most in the first line.
        first_line = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {d: first_line.count(d) for d in (";", ",", "\t")}
        best = max(counts, key=counts.get)  # type: ignore[arg-type]
        return best if counts[best] > 0 else ","


def read_csv(data: bytes, max_rows: int) -> TabularData:
    """Parse CSV bytes into headers + rows.

    Raises:
        ValidationError: empty file, undecodable content or too many rows.
    """
    text = _decode_csv(data)
    delimiter = _sniff_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = [row for row in reader if any(str(c).strip() for c in row)]
    if not all_rows:
        raise ValidationError("O arquivo está vazio.")
    headers = [str(h).strip() for h in all_rows[0]]
    rows = all_rows[1:]
    if len(rows) > max_rows:
        raise ValidationError(f"O arquivo tem {len(rows)} linhas de dados; o limite é {max_rows}.")
    return TabularData(headers=headers, rows=rows)


def read_xlsx(data: bytes, max_rows: int, sheet_name: str | None = None) -> TabularData:
    """Parse XLSX bytes into headers + rows for one sheet.

    Raises:
        ValidationError: corrupt file, unknown sheet, empty sheet or row limit.
    """
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various types for corrupt files
        raise ValidationError("Arquivo XLSX inválido ou corrompido.") from exc

    try:
        sheet_names = list(workbook.sheetnames)
        if sheet_name is not None and sheet_name not in sheet_names:
            raise ValidationError(f"Aba não encontrada: {sheet_name}")
        sheet = workbook[sheet_name] if sheet_name else workbook[sheet_names[0]]

        raw_rows = [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in row)
        ]
    finally:
        workbook.close()

    if not raw_rows:
        raise ValidationError("A aba selecionada está vazia.")
    headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
    rows = raw_rows[1:]
    if len(rows) > max_rows:
        raise ValidationError(f"A aba tem {len(rows)} linhas de dados; o limite é {max_rows}.")
    return TabularData(headers=headers, rows=rows, sheet_names=sheet_names, sheet_name=sheet.title)


def read_json(data: bytes, max_rows: int, sheet_name: str | None = None) -> TabularData:
    """Parse a JSON array of flat objects into headers + rows.

    Every object's keys the union, in first-seen order, become the header row;
    a row missing a key gets ``None`` in that column (never a synthesized
    value — ``None`` reaches the same "missing" pipeline empty CSV cells do).

    Raises:
        ValidationError: not valid JSON, not a list, empty, an element is not
            an object, or too many rows.
    """
    import json

    try:
        payload = json.loads(data.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValidationError("JSON inválido ou não decodificável em UTF-8.") from exc

    if not isinstance(payload, list):
        raise ValidationError("O JSON precisa ser uma lista de objetos.")
    if not payload:
        raise ValidationError("O arquivo está vazio.")
    if not all(isinstance(item, dict) for item in payload):
        raise ValidationError("Cada elemento da lista precisa ser um objeto.")
    if len(payload) > max_rows:
        raise ValidationError(
            f"O arquivo tem {len(payload)} linhas de dados; o limite é {max_rows}."
        )

    headers: list[str] = []
    for item in payload:
        for key in item:
            if key not in headers:
                headers.append(key)
    rows = [[item.get(h) for h in headers] for item in payload]
    return TabularData(headers=headers, rows=rows)


def read_sqlite(data: bytes, max_rows: int, sheet_name: str | None = None) -> TabularData:
    """Parse one table of an uploaded SQLite file into headers + rows.

    ``sheet_name`` (reused from the XLSX contract) selects the table by name;
    with no name given, the first user table (excluding SQLite's own
    ``sqlite_%`` internal tables) is used, mirroring how ``read_xlsx`` defaults
    to the first sheet. ``sqlite3`` needs a real file path, not raw bytes, so
    the upload is written to a temporary file for the duration of the read.

    Raises:
        ValidationError: not a valid SQLite file, no user tables, unknown
            table name, empty table or row limit exceeded.
    """
    import sqlite3
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "upload.sqlite"
        db_path.write_bytes(data)
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = None
            table_names = [
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' "
                    "ORDER BY name"
                ).fetchall()
            ]
            if not table_names:
                raise ValidationError("O arquivo SQLite não tem nenhuma tabela.")
            if sheet_name is not None and sheet_name not in table_names:
                raise ValidationError(f"Tabela não encontrada: {sheet_name}")
            table = sheet_name or table_names[0]

            cursor = conn.execute(f'SELECT * FROM "{table}"')
            headers = [col[0] for col in cursor.description]
            raw_rows = [list(row) for row in cursor.fetchall()]
        except sqlite3.DatabaseError as exc:
            raise ValidationError("Arquivo SQLite inválido ou corrompido.") from exc
        finally:
            conn.close()

    if not raw_rows:
        raise ValidationError("A tabela selecionada está vazia.")
    if len(raw_rows) > max_rows:
        raise ValidationError(
            f"A tabela tem {len(raw_rows)} linhas de dados; o limite é {max_rows}."
        )
    return TabularData(headers=headers, rows=raw_rows, sheet_names=table_names, sheet_name=table)


def read_tabular(
    data: bytes, file_format: str, max_rows: int, sheet_name: str | None = None
) -> TabularData:
    """Dispatch to the reader for ``file_format`` ("csv" | "xlsx" | "json" | "sqlite")."""
    if file_format == "csv":
        return read_csv(data, max_rows)
    if file_format == "xlsx":
        return read_xlsx(data, max_rows, sheet_name)
    if file_format == "json":
        return read_json(data, max_rows, sheet_name)
    if file_format == "sqlite":
        return read_sqlite(data, max_rows, sheet_name)
    raise ValidationError(f"Formato não suportado: {file_format}")
